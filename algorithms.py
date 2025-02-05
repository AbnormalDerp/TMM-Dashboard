import pandas as pd
from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import tempfile
import os


def process_excel(input_file, output_file, start_date, end_date, include_course_types, assets_file, columns_to_keep, rsaf_laptops, a380_laptops,
                  cannot_assign_laptops, cannot_assign_ipads, customers_to_exclude
                  ):
    try:
        # Load the course data
        df = pd.read_excel(input_file)
        
        

        # Remove rows based on conditions
        df = df[df['Course Nature Code'] != 'dry']
        
        # Convert 'From' and 'To' columns to datetime
        df['From'] = pd.to_datetime(df['From']).dt.date
        df['To'] = pd.to_datetime(df['To']).dt.date
        
        # Filter based on start and end date
        start_date = pd.to_datetime(start_date).date()
        end_date = pd.to_datetime(end_date).date()
        df = df[(df['From'] >= start_date) & (df['From'] <= end_date)]
        
        # Keep the necessary columns, including 'Course Type' for filtering
        
        df = df[columns_to_keep]
        
        # Remove the original 'Staff ID' column
        df.drop('Staff ID', axis=1, inplace=True)
        
        # Filter for courses that are in the include list
        df = df[df['Course Type'].isin(include_course_types)]

        # Remove rows where the Customer is in the exclusion list
        df = df[~df['Customer'].isin(customers_to_exclude)]
        
        # Sort data by 'From' date and 'Course'
        df = df.sort_values(by=['From', 'Course'])

        # Load the assets file (CSV)
        assets_df = pd.read_csv(assets_file)

        # Filter only relevant columns (Asset ID, Location, FSA)
        assets_df = assets_df[['Asset ID', 'Location', 'FSA']]

        # Sort the asset data by FSA in descending order and then by Asset ID
        assets_df = assets_df.sort_values(by=['FSA', 'Asset ID'], ascending=[False, True])

        # Exclude laptops with no FSA or an FSA of 'NIL'
        laptops_df = assets_df[
            (assets_df['Location'] == 'M01-13') &
            (assets_df['Asset ID'].str.startswith('L')) &
            (assets_df['FSA'].notna()) & (assets_df['FSA'] != 'NIL')
        ]

        

        
        # Extract iPads from the assets file
        ipads = assets_df[
            (assets_df['Location'] == 'M01-13') & 
            (assets_df['Asset ID'].str.startswith('AIP'))
        ]['Asset ID'].tolist()

        # Debug: Check the extracted iPads
        

        # Exclude restricted iPads
        eligible_ipads = [ipad for ipad in ipads if ipad not in cannot_assign_ipads]

        # Debug: Check eligible iPads after filtering
       

        # Prepare the new columns for laptops and iPads
        df['Staff ID(Lenovo Yoga)'] = None
        df['Staff ID(Apple iPad)'] = None
        df['FSA'] = None

        # Assign laptops based on specific conditions
        for i, row in df.iterrows():
            # Determine the filtered laptops based on conditions
            if row['Customer'] == '99Y':
                # Assign only from RSAF laptops
                filtered_laptops = laptops_df[laptops_df['Asset ID'].isin(rsaf_laptops)]
            elif row['Course Type'].startswith('L') and row['Customer'] == 'SIA':
                # Assign only from A380 laptops
                filtered_laptops = laptops_df[laptops_df['Asset ID'].isin(a380_laptops)]
            else:
                # Assign from laptops not in restricted lists
                filtered_laptops = laptops_df[
                    ~laptops_df['Asset ID'].isin(rsaf_laptops) & 
                    ~laptops_df['Asset ID'].isin(a380_laptops) & 
                    ~laptops_df['Asset ID'].isin(cannot_assign_laptops)
                ]

            # Assign the first available laptop that hasn't been assigned yet
            for _, laptop_row in filtered_laptops.iterrows():
                laptop = laptop_row['Asset ID']
                if laptop not in df['Staff ID(Lenovo Yoga)'].values:  # Ensure no duplication
                    df.at[i, 'Staff ID(Lenovo Yoga)'] = laptop
                    df.at[i, 'FSA'] = laptop_row['FSA']  # Assign the associated FSA
                    break

        ipad_index = 0

        # Exclude restricted iPads
        eligible_ipads = [ipad for ipad in ipads if ipad not in cannot_assign_ipads]
        

        for i, row in df.iterrows():
            # Check if the course type qualifies for an iPad
            if row['Course Type'][0] in ['E', 'G'] and row['Customer'] != '99Y':
                
                if ipad_index < len(eligible_ipads):
                    # Assign the next available eligible iPad
                    df.at[i, 'Staff ID(Apple iPad)'] = eligible_ipads[ipad_index]
                    
                    ipad_index += 1
                else:
                    print(f"No more iPads available for Row {i}")  # Debug: Log when no iPads remain


        # Add the FSA values to the new file, based on Asset ID
        for i, row in df.iterrows():
            if row['Staff ID(Lenovo Yoga)'] is not None:
                laptop_fsa = assets_df[assets_df['Asset ID'] == row['Staff ID(Lenovo Yoga)']]['FSA'].values
                if laptop_fsa:
                    df.at[i, 'FSA'] = laptop_fsa[0]

        # Remove the 'Course Type' column
        df.drop(columns=['Course Type'], inplace=True)

        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Write headers
        for col_num, value in enumerate(df.columns.values, 1):
            ws.cell(row=1, column=col_num, value=value)

        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in [2, 3]:  # 'From' and 'To' columns
                    cell.value = datetime.strptime(str(cell.value), '%Y-%m-%d').strftime('%d-%b-%y')

        # Define color map for the courses
        courses = df['Course'].unique()
        color_map = {'Light Gray': 'D9D9D9', 'Orange': 'FCE4D6'}
        course_colors = {course: color_map['Light Gray'] if i % 2 == 0 else color_map['Orange'] for i, course in enumerate(courses)}

        # Apply colors based on the course
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            course_value = row[0].value  # Assuming 'Course' is the first column
            if course_value in course_colors:
                fill = PatternFill(start_color=course_colors[course_value], end_color=course_colors[course_value], fill_type="solid")
                for cell in row:
                    cell.fill = fill

        # Auto-resize columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Create a temporary file and save the workbook there
        temp_dir = tempfile.gettempdir()  # Get temporary directory path
        temp_file_path = os.path.join(temp_dir, f'{output_file}.xlsx')  # Define temp file path

        wb.save(temp_file_path)  # Save to temp file

        return temp_file_path  # Return path of the saved file for download


    except Exception as e:
        print("Error")


def process_overdue_devices_with_save(excel_file, csv_file, OD_Days, output_file="overdue_devices"):
    """
    Processes the Excel and CSV files to identify overdue devices and save the results to an Excel file.
    """
    
    # Ensure OD_Days is an integer
    if isinstance(OD_Days, str):
        try:
            OD_Days = int(OD_Days)  # Convert to integer if it's a string
        except ValueError:
            print(f"Invalid OD_Days value: {OD_Days}")
            return
    elif isinstance(OD_Days, list):
        if len(OD_Days) > 0 and isinstance(OD_Days[0], int):
            OD_Days = OD_Days[0]  # If OD_Days is a list, use the first element
        else:
            print(f"Invalid OD_Days list value: {OD_Days}")
            return

    # Load Excel and CSV files
    excel_df = pd.read_excel(excel_file)
    csv_df = pd.read_csv(csv_file)

    # Normalize column names
    excel_df.columns = excel_df.columns.str.strip().str.lower()
    csv_df.columns = csv_df.columns.str.strip().str.lower()

    # Check for required columns
    required_columns_excel = {'course', 'from', 'to', 'trainee firstname', 'trainee lastname',
                              'course type name', 'seat number', 'customer', 'customer name'}
    required_columns_csv = {'location', 'asset id'}

    missing_excel = required_columns_excel - set(excel_df.columns)
    missing_csv = required_columns_csv - set(csv_df.columns)

    if missing_excel:
        #print(f"Missing columns in Excel: {missing_excel}")
        return
    if missing_csv:
        #print(f"Missing columns in CSV: {missing_csv}")
        return

    overdue_data = []

    # Iterate through each row in the Excel workbook
    for _, excel_row in excel_df.iterrows():
        course = excel_row['course']
        from_date = excel_row['from']
        trainee_firstname = excel_row['trainee firstname']
        trainee_lastname = excel_row['trainee lastname']
        course_type_name = excel_row['course type name']
        seat_number = excel_row['seat number']
        customer = excel_row['customer']
        customer_name = excel_row['customer name']

        # Debugging: Check if the course exists in CSV
        if course not in csv_df['location'].values:
            #print(f"Course {course} not found in CSV.")
            continue

        # Find all instances of the 'Trainee Firstname' and 'Trainee Lastname' in the workbook
        trainee_matches = excel_df[(
            excel_df['trainee firstname'] == trainee_firstname) & 
            (excel_df['trainee lastname'] == trainee_lastname)
        ]

        # Extract the 'To' dates and find the latest date
        to_dates = []

        # If there are no trainee matches, take the 'To' date from the current row
        if trainee_matches.empty:
            #print(f"No trainee matches for {trainee_firstname} {trainee_lastname} in course {course}. Using the 'To' date from the course.")
            to_dates.append(excel_row['to'])
        else:
            for _, trainee_row in trainee_matches.iterrows():
                to_date = trainee_row['to']
                #print(f"Processing trainee {trainee_row['trainee firstname']} {trainee_row['trainee lastname']}, To Date: {to_date}")

                # Handle both str and Timestamp types
                if isinstance(to_date, pd.Timestamp):
                    to_date = to_date.to_pydatetime()
                elif isinstance(to_date, str):
                    to_date = datetime.strptime(to_date, '%d-%b-%y')

                to_dates.append(to_date)

        if not to_dates:
            #print(f"No 'To' dates found for course {course}")
            continue

        # Use the latest "To" date for overdue check
        latest_to_date = max(to_dates)
        #print(f"Latest To Date for course {course}: {latest_to_date}")

        # If 'From' date and 'To' date are the same, treat them as the same date for overdue logic
        if from_date == latest_to_date:
            print(f"From date and To date are the same for course {course}, treating them as {from_date}")
            latest_to_date = from_date

        current_date = datetime.now()
        #print(f"Current Date: {current_date}")
        #print(f"Checking if course {course} is overdue, comparing to {latest_to_date + timedelta(days=OD_Days)}")

        # Check if the current date is more than OD_Days after the latest 'To' date
        if current_date > latest_to_date + timedelta(days=OD_Days):
            #print(f"Course {course} is overdue.")
            # Get asset IDs for the course
            course_assets = csv_df[csv_df['location'] == course]['asset id'].tolist()
            asset_ids_l = [aid for aid in course_assets if aid.startswith('L')]
            asset_ids_a = [aid for aid in course_assets if aid.startswith('A')]

            # If there are laptops, process them, even if there are no iPads
            if asset_ids_l:
                for laptop_id in asset_ids_l:
                    if asset_ids_a:  # If iPads are available, pair them with laptops
                        ipad_id = asset_ids_a.pop(0)  # Use the first available iPad
                        overdue_data.append({
                            'Course': course,
                            'From': from_date,
                            'To': latest_to_date,
                            'Course Type Name': course_type_name,
                            'Seat Number': '',
                            'Customer': customer,
                            'Customer Name': customer_name,
                            'Trainee Firstname': '',
                            'Trainee Lastname': '',
                            'Staff ID (Lenovo Yoga)': laptop_id,  # Laptop ID
                            'Staff ID (Apple iPad)': ipad_id     # iPad ID
                        })
                    else:  # If no iPads, just record the laptop
                        overdue_data.append({
                            'Course': course,
                            'From': from_date,
                            'To': latest_to_date,
                            'Course Type Name': course_type_name,
                            'Seat Number': '',
                            'Customer': customer,
                            'Customer Name': customer_name,
                            'Trainee Firstname': '',
                            'Trainee Lastname': '',
                            'Staff ID (Lenovo Yoga)': laptop_id,  # Laptop ID
                            'Staff ID (Apple iPad)': None        # No iPad
                        })

            # If there are only iPads (no laptops), process them separately
            if asset_ids_a:
                for ipad_id in asset_ids_a:
                    overdue_data.append({
                        'Course': course,
                        'From': from_date,
                        'To': latest_to_date,
                        'Course Type Name': course_type_name,
                        'Seat Number': '',
                        'Customer': customer,
                        'Customer Name': customer_name,
                        'Trainee Firstname': '',
                        'Trainee Lastname': '',
                        'Staff ID (Lenovo Yoga)': None,  # No Laptop
                        'Staff ID (Apple iPad)': ipad_id  # iPad ID
                    })

    # After gathering all data, we now need to remove duplicates based on 'Staff ID (Lenovo Yoga)'
    if overdue_data:
        #print(f"Overdue data collected: {len(overdue_data)} entries found.")
        overdue_df = pd.DataFrame(overdue_data)
    else:
        print("No overdue devices found.")
        # Create an empty DataFrame with column headers
        columns = ['Course', 'From', 'To', 'Course Type Name', 'Seat Number', 'Customer', 'Customer Name', 
                   'Trainee Firstname', 'Trainee Lastname', 'Staff ID (Lenovo Yoga)', 'Staff ID (Apple iPad)']
        overdue_df = pd.DataFrame(columns=columns)

    # Format 'From' and 'To' columns if there's data
    if not overdue_df.empty:
        overdue_df['From'] = pd.to_datetime(overdue_df['From']).dt.strftime('%d-%b-%y')
        overdue_df['To'] = pd.to_datetime(overdue_df['To']).dt.strftime('%d-%b-%y')

        # Remove duplicate rows based on 'Asset ID (L)', keeping the first occurrence
        overdue_df = overdue_df.drop_duplicates(subset='Staff ID (Lenovo Yoga)', keep='first')

    # Create a temporary file to save the results
    temp_dir = tempfile.gettempdir()  # Get temporary directory path
    temp_file_path = os.path.join(temp_dir, f'{output_file}.xlsx')  # Define temp file path

    with pd.ExcelWriter(temp_file_path, engine='openpyxl') as excel_writer:
        overdue_df.to_excel(excel_writer, index=False)

        # Apply borders after writing the data
        if not overdue_df.empty:
            apply_all_borders(excel_writer.sheets['Sheet1'])

        # Adjust column widths if there's data
        if not overdue_df.empty:
            adjust_column_widths(excel_writer, overdue_df)

        print(f"Overdue devices saved to {temp_file_path} with borders and adjusted column widths.")
    
    return temp_file_path  # Return the path of the saved file




def apply_all_borders(worksheet):
    """
    Applies borders to all cells in the given worksheet.
    """
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Apply borders to all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def adjust_column_widths(excel_writer, df):
    """
    Adjusts the column widths in the Excel file to fit the largest value in each column.
    This mimics the 'double-click to auto-resize' behavior.
    """
    workbook = excel_writer.book
    worksheet = workbook.active

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (A, B, C, ...)
        for cell in col:
            value = str(cell.value)  # Convert to string to handle date and text uniformly
            if len(value) > max_length:
                max_length = len(value)
        adjusted_width = (max_length + 2)  # Add a bit of padding for better appearance
        worksheet.column_dimensions[column].width = adjusted_width


def process_course_data_with_date_filter(assets_file, myteam_file, end_date):
    """
    Process the assets and MyTeam files to find courses, their 'To' dates, and associated assets
    within the range of today to this Thursday (exclusive), ordered by the latest 'To' date.

    Args:
        assets_file (str): Path to the assets CSV file.
        myteam_file (str): Path to the MyTeam Excel file.

    Returns:
        list: A list of strings in the format 'Course - To - Asset ID', with dates in '15 Jan 2025' format.
    """
    try:
        # Get today's date and calculate this Thursday's date
        today = datetime.now().date()
        days_until_friday = (4 - today.weekday() + 7) % 7  # 0 = Monday, 6 = Sunday
        this_friday = today + timedelta(days=days_until_friday)
        this_thursday = this_friday - timedelta(days=1)  # Exclude Friday

        # Load assets file
        assets_df = pd.read_csv(assets_file)

        # Filter locations starting with 'SIN'
        sin_assets = assets_df[assets_df['Location'].str.startswith('SIN', na=False)]

        # Extract relevant data
        location_to_assets = sin_assets.groupby('Location')['Asset ID'].apply(list).to_dict()

        # Load MyTeam file
        myteam_df = pd.read_excel(myteam_file)

        # Initialize results
        results = []

        # Process each SIN location
        for location, asset_ids in location_to_assets.items():
            # Match Location with Course in MyTeam
            course_row = myteam_df[myteam_df['Course'] == location]
            if course_row.empty:
                continue

            # Extract Trainee Code
            trainee_code = course_row['Trainee Code'].iloc[0]

            # Find all instances of Trainee Code
            trainee_rows = myteam_df[myteam_df['Trainee Code'] == trainee_code]

            # Extract the 'To' dates and find the latest date within the range
            to_dates = []

            for _, trainee_row in trainee_rows.iterrows():
                to_date = trainee_row['To']
                # Handle different types for 'To' column
                if isinstance(to_date, pd.Timestamp):
                    to_date = to_date.date()  # Convert to datetime.date
                elif isinstance(to_date, str):
                    to_date = datetime.strptime(to_date, '%d-%b-%y').date()  # Convert to datetime.date
                # Only add to the list if the 'To' date is within the valid range
                if today <= to_date <= end_date:
                    to_dates.append(to_date)

            if not to_dates:
                continue
            # Use the latest "To" date from the list of 'To' dates within the valid range
            latest_to_date = max(to_dates)

            # Format the date as '15 Jan 2025'
            formatted_date = latest_to_date.strftime('%d %b %Y')

            # Compile the result
            asset_ids_str = ', '.join(asset_ids)
            results.append((location, formatted_date, asset_ids_str))

        # Sort results by the earliest 'To' date
        results.sort(key=lambda x: datetime.strptime(x[1], '%d %b %Y'))  # Sort by the date

        # Convert to formatted strings for output
        formatted_results = [f"{course} - {to_date} - {assets}" for course, to_date, assets in results]

        return formatted_results

    except Exception as e:
        print(f"Error processing files: {e}")
        return []


def count_courses_per_month(file_path, include_course_types):
    # Load the Excel file (assuming it contains a sheet with course data)
    df = pd.read_excel(file_path)
    
    # Ensure that 'Course Type' and 'From' columns exist in the dataframe
    if 'Course Type' not in df.columns or 'From' not in df.columns:
        return "Required columns ('Course Type', 'From') are missing from the data."
    
    # Filter courses where 'Course Type' is in the include_course_types list
    filtered_courses = df[df['Course Type'].isin(include_course_types)].copy()
    
    # Convert 'From' to datetime format if it's not already
    filtered_courses['From'] = pd.to_datetime(filtered_courses['From'], errors='coerce')
    
    # Extract the Year-Month period from the 'From' column
    filtered_courses['YearMonth'] = filtered_courses['From'].dt.to_period('M')
    
    # Initialize counters for laptops and iPads per month
    monthly_counts = {}

    # Loop through filtered data and count based on the rules
    for _, row in filtered_courses.iterrows():
        # Extract the relevant data
        course_type = row['Course Type']
        customer = row['Customer']
        year_month = row['YearMonth']
        
        # Initialize the counts for the month if not already in the dictionary
        if year_month not in monthly_counts:
            monthly_counts[year_month] = {'Laptops': 0, 'iPads': 0}
        
        # Count Laptops and iPads based on Course Type and Customer
        if course_type.startswith(('V', 'L')):
            monthly_counts[year_month]['Laptops'] += 1
        elif course_type.startswith(('E', 'G')):
            if customer == '99Y' and course_type.startswith('G'):
                monthly_counts[year_month]['Laptops'] += 1
            else:
                monthly_counts[year_month]['Laptops'] += 1
                monthly_counts[year_month]['iPads'] += 1
    
    # Sort the months and print the results
    results = []  # List to store the results
        
    sorted_months = sorted(monthly_counts.keys())
        
    for month in sorted_months:
        laptops_count = monthly_counts[month]['Laptops']
        ipads_count = monthly_counts[month]['iPads']
            
        # Format the month as "Month Year" using .to_timestamp() and .strftime()
        formatted_month = month.to_timestamp().strftime("%B %Y")
            
        # Add the formatted result to the list
        results.append(f"{formatted_month}: {laptops_count} {ipads_count}")
        
    # Return the list of results
    return results

def process_device_info(myteam_df, assets_df, device_id):
    """
    Processes and retrieves information about a device, including its location,
    course completion percentage, and related assets in the same location.

    Parameters:
    - myteam_df (DataFrame): Data from the myteam Excel workbook.
    - assets_df (DataFrame): Data from the assets CSV file.
    - device_id (str): The Asset ID to look up.

    Returns:
    - dict: A dictionary containing the device information, or an error message.
    """
    assets_df = pd.read_csv(assets_df)
    myteam_df = pd.read_excel(myteam_df)
    # Search for the device in the assets file
    device_row = assets_df[assets_df['Asset ID'] == device_id]

    if device_row.empty:
        return {"error": f"Device ID {device_id} not found in the assets file."}

    location = device_row.iloc[0]['Location']
    if not location.startswith("SIN"):
        return {
            "Asset ID": device_id,
            "Location": location,
            "From": "",
            "To": "",
            "Completion Percentage": 0,
            "Other Asset IDs": []
        }
    
    # Find the corresponding course in the myteam file
    course_row = myteam_df[myteam_df['Course'] == location]
    if course_row.empty:
        return {"error": f"Location {location} not found in the myteam file."}

    # Extract 'From' and 'Trainee Code'
    from_date = course_row.iloc[0]['From']
    trainee_code = course_row.iloc[0]['Trainee Code']

    # Find all 'To' dates for the same Trainee Code
    trainee_rows = myteam_df[myteam_df['Trainee Code'] == trainee_code]
    to_dates = pd.to_datetime(trainee_rows['To'])
    max_to_date = to_dates.max()

    # Calculate completion percentage
    today = datetime.now()
    from_date = pd.to_datetime(from_date)
    completion_percentage = ((today - from_date).days / (max_to_date - from_date).days) * 100
    if completion_percentage > 100:
        completion_percentage = 100
    # Find other Asset IDs with the same location
    same_location_assets = assets_df[assets_df['Location'] == location]['Asset ID'].tolist()
    same_location_assets.remove(device_id)  # Exclude the input device ID

    # Return the result as a dictionary
    return {
        "Asset ID": device_id,
        "Location": location,
        "From": from_date.strftime('%d %b %Y'),
        "To": max_to_date.strftime('%d %b %Y'),
        "Completion Percentage": round(completion_percentage, 2),
        "Other Asset IDs": same_location_assets
    }

def count_fleet_per_month(file_path, include_course_types):
    # Load the Excel file (assuming it contains a sheet with course data)
    df = pd.read_excel(file_path)
    
    # Ensure that 'Course Type' and 'From' columns exist in the dataframe
    if 'Course Type' not in df.columns or 'From' not in df.columns:
        return "Required columns ('Course Type', 'From') are missing from the data."
    
    # Filter courses where 'Course Type' is in the include_course_types list
    filtered_courses = df[df['Course Type'].isin(include_course_types)].copy()
    
    # Convert 'From' to datetime format if it's not already
    filtered_courses['From'] = pd.to_datetime(filtered_courses['From'], errors='coerce')
    
    # Extract the Year-Month period from the 'From' column
    filtered_courses['YearMonth'] = filtered_courses['From'].dt.to_period('M')
    
    # Initialize counters for A320, A330, A350, and A380 per month
    monthly_counts = {}

    # Loop through filtered data and count based on the rules
    for _, row in filtered_courses.iterrows():
        # Extract the relevant data
        course_type = row['Course Type']
        customer = row['Customer']
        year_month = row['YearMonth']
        
        # Initialize the counts for the month if not already in the dictionary
        if year_month not in monthly_counts:
            monthly_counts[year_month] = {'A320': 0, 'A330': 0, 'A350': 0, 'A380': 0}
        
        # Count based on the Course Type first letter
        if course_type.startswith('E'):
            monthly_counts[year_month]['A320'] += 1
        elif course_type.startswith('G'):
            monthly_counts[year_month]['A330'] += 1
        elif course_type.startswith('V'):
            monthly_counts[year_month]['A350'] += 1
        elif course_type.startswith('L'):
            monthly_counts[year_month]['A380'] += 1
    
    # Sort the months and prepare results
    results = []  # List to store the results
        
    sorted_months = sorted(monthly_counts.keys())
        
    for month in sorted_months:
        a320_count = monthly_counts[month]['A320']
        a330_count = monthly_counts[month]['A330']
        a350_count = monthly_counts[month]['A350']
        a380_count = monthly_counts[month]['A380']
            
        # Format the month as "Month Year" using .to_timestamp() and .strftime()
        formatted_month = month.to_timestamp().strftime("%B %Y")
            
        # Add the formatted result to the list
        results.append(f"{formatted_month}: A320: {a320_count}, A330: {a330_count}, A350: {a350_count}, A380: {a380_count}")
        
    # Return the list of results
    return results
