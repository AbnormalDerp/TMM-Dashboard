from flask import Flask, render_template, request, jsonify, send_from_directory, Blueprint
import os
import datetime
from algorithms import process_excel  # Assuming the function is imported from another module
import json
import pandas as pd
import openpyxl
import glob
import tempfile

# Create a Blueprint for middle routes
middle_bp = Blueprint('middle', __name__)


file_name = "Start"

uploaded_files = {}

columns_to_keep = ['Course', 'From', 'To', 'Course Type', 'Course Type Name', 'Seat Number', 'Customer', 'Customer Name', 'Trainee Firstname', 'Trainee Lastname', 'Staff ID']

# Load configuration from the JSON file
def load_config():
    try:
        with open('config.json', 'r') as file:
            config = json.load(file)
        return config
    except Exception as e:
        return None
    
def save_config(updated_config):
    try:
        with open('config.json', 'w') as file:
            json.dump(updated_config, file, indent=4)
        return True
    except Exception as e:
        return False

# Load config data
config = load_config()

# Check if config is loaded properly
if config is None:
    exit()

# Extract values from the config dictionary
rsaf_laptops = config["rsaf_laptops"]
a380_laptops = config["a380_laptops"]
cannot_assign_laptops = config["cannot_assign_laptops"]
cannot_assign_ipads = config["cannot_assign_ipads"]
include_course_types = config["include_course_types"]
customers_to_exclude = config["customers_to_exclude"]
OD_Days = config["OD_Days"]

@middle_bp.route('/')
def index():
    # Check for existing files in the 'uploads' directory
    myteam_file_detected = False
    assets_file_detected = False

    # Check for files that match the naming criteria
    for file in os.listdir('uploads'):
        if file.startswith('SIN') and file.endswith('.xlsx'):
            uploaded_files['myteam'] = file
            myteam_file_detected = True
        if file.startswith('assets') and file.endswith('.csv'):
            uploaded_files['assets'] = file
            assets_file_detected = True

    # Pass the status to the frontend
    return render_template(
        'middle.html',
        myteam_file_detected=myteam_file_detected,
        assets_file_detected=assets_file_detected
    )

# Add a route for the settings button
@middle_bp.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'GET':
        config = load_config()
        # Send the current configuration to the frontend
        return jsonify(config)

    elif request.method == 'POST':
        # Update configuration with data from the frontend
        updated_config = request.get_json()
        if save_config(updated_config):
            # Update global variables
            global rsaf_laptops, a380_laptops, cannot_assign_laptops, cannot_assign_ipads, include_course_types, customers_to_exclude, OD_Days
            rsaf_laptops = updated_config["rsaf_laptops"]
            a380_laptops = updated_config["a380_laptops"]
            cannot_assign_laptops = updated_config["cannot_assign_laptops"]
            cannot_assign_ipads = updated_config["cannot_assign_ipads"]
            include_course_types = updated_config["include_course_types"]
            customers_to_exclude = updated_config["customers_to_exclude"]
            OD_Days = updated_config["OD_Days"]

            # Return the updated config to the frontend
            return jsonify(updated_config)

        else:
            return jsonify({"error": "Failed to save settings."}), 500


@middle_bp.route('/upload-myteam', methods=['POST'])
def upload_myteam():
    file = request.files.get('file')
    if file:
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)
        uploaded_files['myteam'] = file.filename
        return jsonify({"message": "MyTeam file uploaded successfully!", "filename": file.filename})
    return jsonify({"error": "No file uploaded"}), 400

@middle_bp.route('/upload-assets', methods=['POST'])
def upload_assets():
    file = request.files.get('file')
    if file:
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)
        uploaded_files['assets'] = file.filename
        return jsonify({"message": "Assets file uploaded successfully!", "filename": file.filename})
    return jsonify({"error": "No file uploaded"}), 400

@middle_bp.route('/generate', methods=['POST'])
def generate():
    global file_name
    # Get input data from the frontend
    data = request.get_json()
    start_date = data.get('start_date')
    end_date = data.get('end_date')

    # Ensure both start and end dates are provided
    if not start_date or not end_date:
        return jsonify({"error": "Start date and end date are required."}), 400

    # Parse start date for the output file name
    start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    output_filename = f"{start_date_obj.strftime('%d %b %Y')}.xlsx"  # Use the start date for the filename

    # Retrieve the uploaded files
    myteam_file = os.path.join('uploads', uploaded_files.get('myteam', ''))
    assets_file = os.path.join('uploads', uploaded_files.get('assets', ''))

    # Ensure both files are uploaded
    if not myteam_file or not assets_file:
        return jsonify({"error": "Both MyTeam and Assets files are required."}), 400
    
    file_name = output_filename

    # Ensure the file is removed if it already exists in the temp directory
    temp_dir = tempfile.gettempdir()
    temp_file_path = os.path.join(temp_dir, output_filename)
    
    if os.path.exists(temp_file_path):
        os.remove(temp_file_path)

    # Process the Excel file and save it to the same directory as temp directory
    temp_file_path = process_excel(
        input_file=myteam_file,
        output_file=output_filename,
        start_date=start_date,
        end_date=end_date,
        include_course_types=include_course_types,
        assets_file=assets_file,
        columns_to_keep=columns_to_keep,
        rsaf_laptops=rsaf_laptops,
        a380_laptops=a380_laptops,
        cannot_assign_laptops=cannot_assign_laptops,
        cannot_assign_ipads=cannot_assign_ipads,
        customers_to_exclude=customers_to_exclude
    )

    # Save the file in the temp directory, not the project folder
    output_file_path = os.path.join(temp_dir, output_filename)
    os.rename(temp_file_path, output_file_path)  # Move the file to the temp directory
    try:
        # Load the workbook and sheet
        wb = openpyxl.load_workbook(output_file_path)
        sheet = wb.active

        # Create a list to hold the table rows
        table_data = []
        max_column_widths = {col: 0 for col in range(1, sheet.max_column + 1)}  # Track the max width for each column

        for row in sheet.iter_rows(min_row=2):  # Start from row 2 to avoid duplicating the header
            row_data = []
            for col_index, cell in enumerate(row, start=1):
                value = cell.value if cell.value is not None else ""
                color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                row_data.append((value, color))

            table_data.append(row_data)

        # Column widths based on your provided data
        column_widths = {
            "Course": 14.36,
            "From": 10.36,
            "To": 10.36,
            "Course Type Name": 40.36,
            "Seat Number": 12.36,
            "Customer": 9.36,
            "Customer Name": 25.36,
            "Trainee Firstname": 18.36,
            "Trainee Lastname": 17.36,
            "Staff ID(Lenovo Yoga)": 22.36,
            "Staff ID(Apple iPad)": 21.26,
            "FSA": 5.36
        }

        # Generate HTML table for the frontend
        html_table = '<table class="excel-table">'

        # Add headers with white background and black text (Only once)
        html_table += '<thead><tr>'
        for col_index in range(sheet.max_column):
            header_value = sheet.cell(row=1, column=col_index + 1).value
            html_table += f'<th style="background-color: white; color: black; border: 1px solid #ddd;">{header_value}</th>'
        html_table += '</tr></thead>'

        html_table += '<tbody>'
        for row_data in table_data:
            html_table += '<tr>'
            for value, color in row_data:
                style = f'background-color: #{color[2:]}' if color else ''
                html_table += f'<td style="{style}">{value}</td>'
            html_table += '</tr>'
        html_table += '</tbody></table>'

        # Send back the output file link and HTML table
        return jsonify({
            "message": "Process complete.",
            "output_file": f'/download/{output_filename}',  # Provide path for download
            "html_table": html_table,
            "column_widths": column_widths
        })

    except Exception as e:
        return jsonify({"error": f"An error occurred while reading the generated file: {str(e)}"}), 500



@middle_bp.route('/download/<filename>')
def download_file(filename):
    print(filename)
    temp_dir = tempfile.gettempdir()  # Directory where the file is saved
    if os.path.exists(os.path.join(temp_dir, filename)):
        return send_from_directory(temp_dir, filename, as_attachment=True)
    else:
        return jsonify({"error": "File not found."}), 404




if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    middle_bp.run(debug=True)
